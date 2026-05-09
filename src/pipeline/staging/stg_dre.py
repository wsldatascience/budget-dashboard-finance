"""
Staging layer: DRE sheet
Grain: one row per conta contábil (leaf rows only, those with cod_conta)

DRE structure (0-indexed columns):
  col 0 = cod_conta (numeric float, only on leaf rows)
  col 2 = nome_conta
  col 4 = Janeiro orçado
  col 6 = Fevereiro orçado
  col 8 = Março orçado
  col 10 = Acumulado Q1 orçado
"""
import pandas as pd
import logging

from src.config import KNOWN_UNBUDGETED_ACCOUNTS, ROLLUP_ACCOUNTS

logger = logging.getLogger(__name__)

DATA_START_ROW = 5   # first data row (0-indexed) after the 2-row header block


def read(filepath: str, sheet_name: str = "DRE") -> pd.DataFrame:
    raw = pd.read_excel(filepath, sheet_name=sheet_name, header=None)
    logger.info("DRE raw shape: %s", raw.shape)

    df = raw.iloc[DATA_START_ROW:].copy().reset_index(drop=True)

    result = pd.DataFrame({
        "cod_conta":    df.iloc[:, 0],
        "nome_conta":   df.iloc[:, 2],
        "jan_orcado":   pd.to_numeric(df.iloc[:, 4], errors="coerce").fillna(0),
        "fev_orcado":   pd.to_numeric(df.iloc[:, 6], errors="coerce").fillna(0),
        "mar_orcado":   pd.to_numeric(df.iloc[:, 8], errors="coerce").fillna(0),
        "acumulado_q1": pd.to_numeric(df.iloc[:, 10], errors="coerce").fillna(0),
    })

    # Keep only leaf rows (those with an account code)
    result = result.dropna(subset=["cod_conta"]).copy()
    result["cod_conta"] = result["cod_conta"].apply(_normalize_code).astype(str)
    result["nome_conta"] = result["nome_conta"].astype(str).str.strip()

    # Fix accounts whose monthly values are 0 due to unresolved Excel formulas.
    result = _patch_formula_accounts(df, result, filepath, sheet_name)

    recalculated = result["jan_orcado"] + result["fev_orcado"] + result["mar_orcado"]
    # Excel acumulado_q1 == 0 typically means the SUM formula was not cached
    # (legit case after patch). Patched accounts also legitimately diverge —
    # the divergence is the patched delta itself.
    patched_codes = result.attrs.get("patched_codes", set())
    real_divergence = result[
        (abs(recalculated - result["acumulado_q1"]) > 0.02)
        & (result["acumulado_q1"].abs() > 0.02)
        & (~result["cod_conta"].isin(patched_codes))
    ]
    if not real_divergence.empty:
        logger.warning(
            "acumulado_q1 diverges from jan+fev+mar for %d non-patched "
            "account(s) with non-zero Excel total: %s — using recalculated "
            "sum. Verify for non-additive Excel formulas.",
            len(real_divergence), real_divergence["cod_conta"].tolist(),
        )
    result["acumulado_q1"] = recalculated
    result = _inject_known_unbudgeted(result)

    logger.info("DRE leaf rows extracted: %d", len(result))
    _validate(result)
    return result


def _patch_formula_accounts(
    df: pd.DataFrame,
    result: pd.DataFrame,
    filepath: str,
    sheet_name: str,
) -> pd.DataFrame:
    """
    Patch leaf accounts whose Excel cell contains a SUM formula that is not
    evaluated when reading with data_only=True (cached value missing).

    Two patch strategies, applied conservatively:

    1. Fixed-monthly budgets (fev == mar but jan = 0): replicate fev to jan.

    2. Roll-up accounts (declared in ROLLUP_ACCOUNTS config): the leaf is a
       SUM over decorative breakdown rows in the same Excel section. Patch
       leaf = nearest preceding section subtotal (whichever outline_level
       is below the leaf's outline_level). This avoids the previous bug
       that walked back through ancestor subtotals and double-counted them.

    Accounts that are all-zero AND not in ROLLUP_ACCOUNTS are LEFT AS ZERO.
    These are intentional budget gaps (the formula did evaluate to 0). The
    previous heuristic walk-back inflated 320101 (Medicamentos) to R$13M
    and 360101 (Juros) to R$1.27M by summing parent subtotals — strictly
    worse than no patch.
    """
    import openpyxl
    result = result.copy()

    # Build outline_level + value lookup once from the workbook.
    # Map: excel_row_num → (outline_level, jan, fev, mar, q1, has_code)
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet_name]
    row_meta: dict[int, dict] = {}
    for row_num, raw in enumerate(
        ws.iter_rows(min_row=DATA_START_ROW + 2, max_row=ws.max_row, values_only=True),
        start=DATA_START_ROW + 2,
    ):
        ol = ws.row_dimensions[row_num].outline_level
        row_meta[row_num] = {
            "ol":   ol,
            "cod":  raw[0],
            "jan":  raw[4]  if len(raw) > 4  else None,
            "fev":  raw[6]  if len(raw) > 6  else None,
            "mar":  raw[8]  if len(raw) > 8  else None,
            "q1":   raw[10] if len(raw) > 10 else None,
        }

    cod_to_row: dict[str, int] = {}
    for rn, m in row_meta.items():
        if m["cod"] is not None:
            try:
                cod_to_row[_normalize_code(m["cod"])] = rn
            except Exception:
                pass

    patched_codes: set[str] = set()
    for idx, row in result.iterrows():
        cod = row["cod_conta"]

        # Strategy 1: jan missing but fev == mar (fixed monthly budget)
        if (row["jan_orcado"] == 0 and row["fev_orcado"] > 0 and
                abs(row["fev_orcado"] - row["mar_orcado"]) < 0.01):
            result.at[idx, "jan_orcado"] = result.at[idx, "fev_orcado"]
            patched_codes.add(cod)
            logger.info(
                "Patched fixed-monthly account %s: jan=%.2f",
                cod, result.at[idx, "fev_orcado"],
            )
            continue

        all_zero = (row["jan_orcado"] == 0 and row["fev_orcado"] == 0
                    and row["mar_orcado"] == 0)
        if not all_zero:
            continue

        # Strategy 2: declared roll-up account → use nearest preceding
        # section subtotal (outline_level strictly less than leaf's level
        # AND with cached numeric values).
        if cod not in ROLLUP_ACCOUNTS:
            logger.info(
                "Account %s: all-zero in Excel and not in ROLLUP_ACCOUNTS — "
                "leaving as zero (budget intentionally empty)", cod,
            )
            continue

        leaf_row = cod_to_row.get(cod)
        if leaf_row is None:
            logger.warning("ROLLUP account %s: row not found in workbook", cod)
            continue

        leaf_ol = row_meta[leaf_row]["ol"]
        section = None
        for rn in range(leaf_row - 1, DATA_START_ROW + 1, -1):
            m = row_meta.get(rn)
            if m is None:
                continue
            if m["cod"] is not None:
                # Hit another leaf account before finding section — stop
                break
            if m["ol"] >= leaf_ol:
                continue
            if all(v is not None for v in (m["jan"], m["fev"], m["mar"])):
                section = m
                break

        if section is None:
            logger.warning(
                "ROLLUP account %s: no preceding section subtotal found", cod,
            )
            continue

        result.at[idx, "jan_orcado"] = round(float(section["jan"]), 2)
        result.at[idx, "fev_orcado"] = round(float(section["fev"]), 2)
        result.at[idx, "mar_orcado"] = round(float(section["mar"]), 2)
        patched_codes.add(cod)
        logger.info(
            "Patched ROLLUP account %s from section subtotal: "
            "jan=%.2f fev=%.2f mar=%.2f",
            cod, section["jan"], section["fev"], section["mar"],
        )

    # Annotate so the divergence check can skip accounts we intentionally patched
    result.attrs["patched_codes"] = patched_codes
    return result


def _inject_known_unbudgeted(df: pd.DataFrame) -> pd.DataFrame:
    """
    Inject P&L accounts that carry realized expenses but have no DRE budget line.
    Added with orcado=0 so the mart tracks them as unplanned spending (WARNING)
    rather than flagging them as orphan accounts (FAIL).
    """
    existing = set(df["cod_conta"])
    rows = []
    for cod, nome in KNOWN_UNBUDGETED_ACCOUNTS.items():
        if cod not in existing:
            rows.append({
                "cod_conta": cod, "nome_conta": nome,
                "jan_orcado": 0.0, "fev_orcado": 0.0, "mar_orcado": 0.0,
                "acumulado_q1": 0.0,
            })
            logger.info("Injected known unbudgeted account %s: %s", cod, nome)
    if rows:
        df = pd.concat([df, pd.DataFrame(rows)], ignore_index=True)
    return df


def _normalize_code(val) -> str:
    """Convert float code (310101.0) to string ('310101')."""
    try:
        return str(int(float(val)))
    except (ValueError, TypeError):
        return str(val).strip()


def _validate(df: pd.DataFrame) -> None:
    if df.empty:
        raise ValueError("DRE staging returned empty DataFrame")
    dupes = df[df["cod_conta"].duplicated(keep=False)]
    if not dupes.empty:
        logger.warning("Duplicate cod_conta in DRE: %s", dupes["cod_conta"].tolist())
