"""
Generates dashboard/public/dashboard_data.json consumed by the React dashboard.

Produces:
  - summary KPIs
  - monthly breakdown (Jan / Fev / Mar) — orçado and realizado
  - by_group aggregation
  - full conta list
  - quality alerts
"""
import json
import logging
import sys
from pathlib import Path

import pandas as pd

_ROOT = Path(__file__).resolve().parent.parent
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

from src.config import (
    EMPRESA_FILTER, PERIODO_LABEL, GROUP_PREFIXES, derive_group,
    RECEITA_COD_CONTA, KNOWN_UNBUDGETED_ACCOUNTS,
)
from src.utils import compute_variance_pct, file_sha256

ROOT       = Path(__file__).resolve().parent.parent
DATA_FILE  = ROOT / "data" / "teste_budget.xlsx"
OUT_FILE   = ROOT / "dashboard" / "public" / "dashboard_data.json"

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
logger = logging.getLogger(__name__)

MONTHS = [
    {"key": "jan", "col": 4, "label": "Janeiro"},
    {"key": "fev", "col": 6, "label": "Fevereiro"},
    {"key": "mar", "col": 8, "label": "Março"},
]



def build() -> dict:
    from src.pipeline.staging import stg_dre as _stg_dre_mod
    from src.pipeline.staging import stg_despesas as _stg_desp_mod

    # ── Staging — read each source once ──────────────────────────────────────
    stg_dre_full = _stg_dre_mod.read(str(DATA_FILE))
    desp         = _stg_desp_mod.read(str(DATA_FILE))

    # Despesas DRE: exclude 31xx revenue/abatimento accounts
    stg_dre_desp = stg_dre_full[~stg_dre_full["cod_conta"].str.startswith("31")].copy()
    stg_dre_desp = stg_dre_desp.rename(columns={
        "jan_orcado": "orc_jan", "fev_orcado": "orc_fev",
        "mar_orcado": "orc_mar", "acumulado_q1": "orc_q1",
    })
    stg_dre_desp["grupo"] = stg_dre_desp["cod_conta"].apply(derive_group)
    dre = stg_dre_desp[["cod_conta", "nome_conta", "grupo", "orc_jan", "orc_fev", "orc_mar", "orc_q1"]].copy()

    cde = desp[desp["empresa"] == EMPRESA_FILTER].copy()
    cde["mes"] = cde["data"].dt.month  # 1=Jan, 2=Fev, 3=Mar

    # Monthly realized per conta
    for m_num, m_key in [(1, "jan"), (2, "fev"), (3, "mar")]:
        monthly = (
            cde[cde["mes"] == m_num]
            .groupby("cod_conta")["valor"]
            .sum()
            .reset_index()
            .rename(columns={"valor": f"rea_{m_key}"})
        )
        dre = dre.merge(monthly, on="cod_conta", how="left")
        dre[f"rea_{m_key}"] = dre[f"rea_{m_key}"].fillna(0)

    dre["rea_q1"] = dre["rea_jan"] + dre["rea_fev"] + dre["rea_mar"]
    dre["var_rs"]  = dre["rea_q1"] - dre["orc_q1"]
    dre["var_pct"] = dre.apply(lambda r: compute_variance_pct(r["rea_q1"], r["orc_q1"]), axis=1)

    # ── Summary KPIs ─────────────────────────────────────────────────────────
    total_orc = round(dre["orc_q1"].sum(), 2)
    total_rea = round(dre["rea_q1"].sum(), 2)
    summary = {
        "total_orcado":    total_orc,
        "total_realizado": total_rea,
        "variacao_rs":     round(total_rea - total_orc, 2),
        "variacao_pct":    round((total_rea - total_orc) / abs(total_orc) * 100, 2) if total_orc else 0,
    }

    # ── Monthly totals ────────────────────────────────────────────────────────
    monthly = []
    for m in MONTHS:
        k = m["key"]
        orc = round(dre[f"orc_{k}"].sum(), 2)
        rea = round(dre[f"rea_{k}"].sum(), 2)
        monthly.append({
            "mes":       m["label"],
            "orcado":    orc,
            "realizado": rea,
            "variacao_rs":  round(rea - orc, 2),
            "variacao_pct": round((rea - orc) / abs(orc) * 100, 2) if orc else 0,
        })

    # ── By group ─────────────────────────────────────────────────────────────
    grp = (
        dre.groupby("grupo")[["orc_q1", "rea_q1"]]
        .sum()
        .reset_index()
        .rename(columns={"orc_q1": "orcado", "rea_q1": "realizado"})
    )
    grp["variacao_rs"]  = (grp["realizado"] - grp["orcado"]).round(2)
    grp["variacao_pct"] = grp.apply(lambda r: compute_variance_pct(r["realizado"], r["orcado"]), axis=1)
    grp = grp.sort_values("variacao_rs")
    by_group = grp.to_dict(orient="records")

    # ── Conta ranking (top variances) ─────────────────────────────────────────
    ranking = (
        dre[dre["orc_q1"] != 0]
        .nlargest(15, "var_rs")[["cod_conta", "nome_conta", "grupo", "orc_q1", "rea_q1", "var_rs", "var_pct"]]
        .rename(columns={"orc_q1": "orcado", "rea_q1": "realizado",
                         "var_rs": "variacao_rs", "var_pct": "variacao_pct"})
    )
    ranking_bottom = (
        dre[dre["orc_q1"] != 0]
        .nsmallest(15, "var_rs")[["cod_conta", "nome_conta", "grupo", "orc_q1", "rea_q1", "var_rs", "var_pct"]]
        .rename(columns={"orc_q1": "orcado", "rea_q1": "realizado",
                         "var_rs": "variacao_rs", "var_pct": "variacao_pct"})
    )

    # ── All contas ────────────────────────────────────────────────────────────
    all_contas = dre[[
        "cod_conta", "nome_conta", "grupo",
        "orc_jan", "orc_fev", "orc_mar", "orc_q1",
        "rea_jan", "rea_fev", "rea_mar", "rea_q1",
        "var_rs", "var_pct",
    ]].copy()
    # Round monetary columns
    money_cols = [c for c in all_contas.columns if c not in ("cod_conta", "nome_conta", "grupo", "var_pct")]
    all_contas[money_cols] = all_contas[money_cols].round(2)

    # ── Quality alerts ────────────────────────────────────────────────────────
    contas_dre = set(dre["cod_conta"].unique())
    cde_agg = cde.groupby(["cod_conta", "nome_conta"])["valor"].sum().reset_index()
    missing = cde_agg[~cde_agg["cod_conta"].isin(contas_dre)].copy()
    missing["valor"] = missing["valor"].round(2)
    missing = missing.sort_values("valor", ascending=False)

    unplanned = dre[(dre["orc_q1"] == 0) & (dre["rea_q1"] > 0)][
        ["cod_conta", "nome_conta", "rea_q1"]
    ].rename(columns={"rea_q1": "valor_realizado"})
    unplanned["valor_realizado"] = unplanned["valor_realizado"].round(2)

    # ── Receita — reuse stg_dre_full already in memory ───────────────────────
    from src.pipeline.staging import stg_receita as _stg_rec
    from src.pipeline.intermediate import int_receita as _int_rec
    from src.pipeline.marts import fct_receita_vs_orcado as _fct_rec

    receita_raw         = _stg_rec.read(str(DATA_FILE))
    receita_by_conta    = _int_rec.build_by_conta(receita_raw)
    receita_by_convenio = _int_rec.build_by_convenio(receita_raw)
    receita_by_grupo    = _int_rec.build_by_grupo(receita_raw)
    mart_receita        = _fct_rec.build(stg_dre_full, receita_by_conta)

    # Receita summary KPIs
    rec_bruta = float(receita_by_conta["valor_realizado"].iloc[0])
    rec_orcado = float(
        mart_receita.loc[mart_receita["cod_conta"] == RECEITA_COD_CONTA, "valor_orcado"].sum()
    )
    rec_deducoes_orc = float(
        mart_receita.loc[mart_receita["cod_conta"] != RECEITA_COD_CONTA, "valor_orcado"].sum()
    )
    rec_liquida_estimada = round(rec_bruta - abs(rec_deducoes_orc), 2)

    receita_summary = {
        "receita_bruta_realizada": round(rec_bruta, 2),
        "receita_bruta_orcada":    round(rec_orcado, 2),
        "variacao_rs":             round(rec_bruta - rec_orcado, 2),
        "variacao_pct":            round((rec_bruta - rec_orcado) / abs(rec_orcado) * 100, 2) if rec_orcado else None,
        "deducoes_orcadas":        round(abs(rec_deducoes_orc), 2),
        "receita_liquida_estimada": rec_liquida_estimada,
        "rea_jan": float(receita_by_conta["rea_jan"].iloc[0]),
        "rea_fev": float(receita_by_conta["rea_fev"].iloc[0]),
        "rea_mar": float(receita_by_conta["rea_mar"].iloc[0]),
    }

    # Receita by convênio (Q1 total per convênio)
    conv_q1 = (
        receita_by_convenio.groupby("convenio")["valor_realizado"]
        .sum()
        .reset_index()
        .sort_values("valor_realizado", ascending=False)
        .rename(columns={"valor_realizado": "total_q1"})
    )
    conv_q1["total_q1"] = conv_q1["total_q1"].round(2)
    conv_q1["pct_receita"] = (conv_q1["total_q1"] / rec_bruta * 100).round(2) if rec_bruta else 0

    # Receita by grupo de produto (Q1 total)
    grupo_q1 = (
        receita_by_grupo.groupby("grupo_produto")["valor_realizado"]
        .sum()
        .reset_index()
        .sort_values("valor_realizado", ascending=False)
        .rename(columns={"valor_realizado": "total_q1"})
    )
    grupo_q1["total_q1"] = grupo_q1["total_q1"].round(2)
    grupo_q1["pct_receita"] = (grupo_q1["total_q1"] / rec_bruta * 100).round(2) if rec_bruta else 0

    # ── Margens & EBITDA ──────────────────────────────────────────────────────
    grp_lookup = {g["grupo"]: g for g in by_group}

    _CSP_GROUPS = {"Honorários Médicos", "Pessoal de Produção", "Insumos Assistenciais"}
    _DOP_GROUPS = {
        "Pessoal Administrativo", "Serviços de Terceiros", "Infraestrutura",
        "Materiais de Consumo", "Utilidades e Serviços",
        "Manutenção e Reparos", "Manutenção de Veículos", "Despesas Diversas",
    }

    def _sum(groups):
        return {
            "orcado":    sum(grp_lookup[g]["orcado"]    for g in groups if g in grp_lookup),
            "realizado": sum(grp_lookup[g]["realizado"] for g in groups if g in grp_lookup),
        }

    csp     = _sum(_CSP_GROUPS)
    desp_op = _sum(_DOP_GROUPS)
    da      = grp_lookup.get("Depreciação e Amortização", {"orcado": 0, "realizado": 0})

    custo_orcado    = csp["orcado"]
    custo_realizado = csp["realizado"]

    rec_bruta_orc = rec_orcado
    rec_bruta_rea = rec_bruta
    abat_orc      = abs(rec_deducoes_orc)
    # Abatimentos realizados: somar mart_receita excluindo 310101
    abat_rea = float(
        mart_receita.loc[mart_receita["cod_conta"] != RECEITA_COD_CONTA, "valor_realizado"].sum()
    )
    rec_liq_orc = rec_bruta_orc - abat_orc
    rec_liq_rea = rec_bruta_rea - abat_rea

    lucro_bruto_orc = rec_liq_orc - custo_orcado
    lucro_bruto_rea = rec_liq_rea - custo_realizado

    ebitda_orc = lucro_bruto_orc - desp_op["orcado"]
    ebitda_rea = lucro_bruto_rea - desp_op["realizado"]

    def _margem(num: float, den: float) -> float | None:
        return round(num / den * 100, 2) if den else None

    # Lucro Operacional = EBITDA − D&A (margem operacional inclui depreciação)
    lucro_op_orc = ebitda_orc - da["orcado"]
    lucro_op_rea = ebitda_rea - da["realizado"]

    margens = {
        "receita_bruta":        {"orcado": round(rec_bruta_orc, 2), "realizado": round(rec_bruta_rea, 2)},
        "abatimentos":          {"orcado": round(abat_orc, 2),     "realizado": round(abat_rea, 2)},
        "receita_liquida":      {"orcado": round(rec_liq_orc, 2),  "realizado": round(rec_liq_rea, 2)},
        "custo_servicos":       {"orcado": round(custo_orcado, 2), "realizado": round(custo_realizado, 2)},
        "lucro_bruto":          {"orcado": round(lucro_bruto_orc, 2), "realizado": round(lucro_bruto_rea, 2)},
        "despesas_operacionais":{"orcado": round(desp_op["orcado"], 2), "realizado": round(desp_op["realizado"], 2)},
        "depreciacao_amortizacao": {"orcado": round(da["orcado"], 2), "realizado": round(da["realizado"], 2)},
        "ebitda":               {"orcado": round(ebitda_orc, 2),   "realizado": round(ebitda_rea, 2)},
        "lucro_operacional":    {"orcado": round(lucro_op_orc, 2), "realizado": round(lucro_op_rea, 2)},
        "margem_bruta_pct":     {"orcado": _margem(lucro_bruto_orc, rec_liq_orc), "realizado": _margem(lucro_bruto_rea, rec_liq_rea)},
        "margem_operacional_pct": {"orcado": _margem(lucro_op_orc, rec_liq_orc),  "realizado": _margem(lucro_op_rea, rec_liq_rea)},
        "margem_ebitda_pct":    {"orcado": _margem(ebitda_orc, rec_liq_orc),      "realizado": _margem(ebitda_rea, rec_liq_rea)},
    }

    # ── EBITDA Pro-forma — ajusta provisões esperadas porém não realizadas ───
    # Contas com gap material entre orçado e realizado por sub-provisionamento:
    #   - Abatimentos s/ serviços (PIS/COFINS/ISS): já capturado em margens.abatimentos
    #     mas o gap é deduzido da receita líquida, não do EBITDA → trata-se em separado
    #   - 330401 Depreciação Patrimonial → Despesa Operacional (entra no EBITDA)
    #   - 320204 13º Salário Produção → Custo (entra no EBITDA via Lucro Bruto)
    #   - 330108 13° Salário Administração → Despesa Operacional (entra no EBITDA)
    #   - 390101 IR PJ / 390102 CSLL → abaixo da linha (NÃO afeta EBITDA, mas
    #     compõe o gap de "EBITDA — IR — CSLL" para comparar com Lucro Líquido)
    # afeta_ebitda=True só para itens que ENTRAM no EBITDA por definição
    # (custos e despesas operacionais). D&A e IR/CSLL ficam abaixo da linha
    # do EBITDA — provisionar elas só desce o Lucro Líquido, não o EBITDA.
    PROFORMA_ACCOUNTS = {
        "320204": {"label": "13º Salário (Produção)",      "afeta_ebitda": True,  "tipo": "provisao_pessoal"},
        "330108": {"label": "13º Salário (Administração)", "afeta_ebitda": True,  "tipo": "provisao_pessoal"},
        "330401": {"label": "Depreciação Patrimonial",     "afeta_ebitda": False, "tipo": "depreciacao"},
        "390101": {"label": "IR Pessoa Jurídica",          "afeta_ebitda": False, "tipo": "imposto"},
        "390102": {"label": "Contribuição Social",         "afeta_ebitda": False, "tipo": "imposto"},
    }
    contas_lookup = dre.set_index("cod_conta")[["nome_conta", "orc_q1", "rea_q1"]].to_dict(orient="index")
    ajustes_lista = []
    gap_ebitda_total = 0.0
    gap_pos_ebitda   = 0.0  # D&A + IR/CSLL — deduzidos só do Lucro Líquido
    for cod, meta in PROFORMA_ACCOUNTS.items():
        info = contas_lookup.get(cod)
        if not info:
            continue
        gap = max(0.0, float(info["orc_q1"]) - float(info["rea_q1"]))
        if gap <= 0:
            continue
        ajustes_lista.append({
            "cod_conta": cod,
            "nome": meta["label"],
            "tipo": meta["tipo"],
            "orcado_q1": round(float(info["orc_q1"]), 2),
            "realizado_q1": round(float(info["rea_q1"]), 2),
            "ajuste_rs": round(gap, 2),
            "afeta_ebitda": meta["afeta_ebitda"],
        })
        if meta["afeta_ebitda"]:
            gap_ebitda_total += gap
        else:
            gap_pos_ebitda += gap

    # Gap de abatimentos (PIS/COFINS/ISS): impacta receita líquida, então
    # impacta o EBITDA pro-forma (custos/desp permanecem, RL cai → EBITDA cai)
    gap_abatimentos = max(0.0, abat_orc - abat_rea)
    if gap_abatimentos > 0:
        ajustes_lista.append({
            "cod_conta": None,
            "nome": "Abatimentos (PIS/COFINS/ISS)",
            "tipo": "tributos_servico",
            "orcado_q1": round(abat_orc, 2),
            "realizado_q1": round(abat_rea, 2),
            "ajuste_rs": round(gap_abatimentos, 2),
            "afeta_ebitda": True,
        })

    rec_liq_proforma = rec_liq_rea - gap_abatimentos
    ebitda_proforma  = ebitda_rea - gap_ebitda_total - gap_abatimentos
    margem_proforma  = _margem(ebitda_proforma, rec_liq_proforma)
    # LL aprox = EBITDA pf − (D&A + IR + CSLL provisionados). Ainda ignora
    # Resultado Financeiro e PSR — sufixo "aprox" deixa essa limitação explícita.
    lucro_liq_proforma = ebitda_proforma - gap_pos_ebitda

    margens["ebitda_proforma"] = {
        "ebitda_realizado":      round(ebitda_rea, 2),
        "ebitda_proforma":       round(ebitda_proforma, 2),
        "ajuste_total_rs":       round(gap_ebitda_total + gap_abatimentos, 2),
        "receita_liquida_proforma": round(rec_liq_proforma, 2),
        "margem_ebitda_realizado_pct": _margem(ebitda_rea, rec_liq_rea),
        "margem_ebitda_proforma_pct":  margem_proforma,
        "lucro_liquido_proforma_aprox": round(lucro_liq_proforma, 2),
        "ajustes": ajustes_lista,
    }

    # ── Concentração de clientes (HHI + Top-N) ───────────────────────────────
    conv_records = conv_q1.to_dict(orient="records")
    pcts = [float(r.get("pct_receita") or 0) for r in conv_records]
    hhi = round(sum(p * p for p in pcts), 1)  # Herfindahl-Hirschman Index
    if hhi >= 2500:
        hhi_classificacao = "Altamente concentrado"
    elif hhi >= 1500:
        hhi_classificacao = "Moderadamente concentrado"
    else:
        hhi_classificacao = "Pouco concentrado"

    top1 = conv_records[0] if conv_records else None
    top3_pct = round(sum(pcts[:3]), 2)
    top5_pct = round(sum(pcts[:5]), 2)
    n_convenios = len(conv_records)

    concentracao = {
        "hhi": hhi,
        "hhi_classificacao": hhi_classificacao,
        "top1_convenio": top1["convenio"] if top1 else None,
        "top1_pct":      round(float(top1["pct_receita"]) if top1 else 0, 2),
        "top1_valor":    round(float(top1["total_q1"])     if top1 else 0, 2),
        "top3_pct":      top3_pct,
        "top5_pct":      top5_pct,
        "n_convenios":   n_convenios,
        "top5":          conv_records[:5],
    }

    # ── DRE full hierarchy ────────────────────────────────────────────────────
    dre_hierarchy = _build_dre_hierarchy(str(DATA_FILE), mart_receita, desp, stg_dre_full)

    return _clean({
        "generated_at": pd.Timestamp.now().isoformat(),
        "input_hash":   file_sha256(DATA_FILE),
        "empresa":  EMPRESA_FILTER,
        "periodo":  PERIODO_LABEL,
        "summary":  summary,
        "monthly":  monthly,
        "by_group": by_group,
        "ranking_above": ranking.to_dict(orient="records"),
        "ranking_below": ranking_bottom.to_dict(orient="records"),
        "all_contas": all_contas.to_dict(orient="records"),
        "quality": {
            "missing_contas": missing.to_dict(orient="records"),
            "unplanned_spending": unplanned.to_dict(orient="records"),
        },
        "receita": {
            "summary":       receita_summary,
            "dre_accounts":  mart_receita.to_dict(orient="records"),
            "by_convenio":   conv_q1.to_dict(orient="records"),
            "by_grupo":      grupo_q1.to_dict(orient="records"),
            "concentracao":  concentracao,
        },
        "margens": margens,
        "dre": dre_hierarchy,
    })


def _build_dre_hierarchy(
    excel_path: str,
    mart_rec_df,
    desp_df: pd.DataFrame,
    stg_dre_df: pd.DataFrame,
) -> list:
    """
    Build the full DRE hierarchy preserving Excel row order.
    Account rows get orçado from the staging module (formula-patched)
    and realized from both marts.
    desp_df / stg_dre_df: already-read DataFrames — no extra Excel reads here.
    """
    import openpyxl

    MAJOR_SECTIONS = {
        "Receita Bruta", "Abatimentos", "Receita L\u00edquida",
        "Custo dos Servi\u00e7os Prestados", "Lucro Bruto",
        "Despesas Operacionais", "EBITDA",
        "Deprecia\u00e7\u00e3o E Amortiza\u00e7\u00e3o", "Lucro Operacional",
        "Resultado Financeiro", "Imposto De Renda e CSLL",
        "Participa\u00e7\u00e3o Sobre Os Resultados", "Lucro L\u00edquido",
    }

    # ── Orçado lookup (formula-patched staging values) ────────────────────────
    orc_lookup: dict = {}
    for _, r in stg_dre_df.iterrows():
        orc_lookup[str(r["cod_conta"])] = {
            "jan": float(r["jan_orcado"]),
            "fev": float(r["fev_orcado"]),
            "mar": float(r["mar_orcado"]),
            "q1":  float(r["jan_orcado"] + r["fev_orcado"] + r["mar_orcado"]),
        }

    # ── Realized lookup ───────────────────────────────────────────────────────
    realized: dict = {}
    for _, r in mart_rec_df.iterrows():
        realized[str(r["cod_conta"])] = float(r.get("valor_realizado", 0) or 0)

    xpto_desp = desp_df[desp_df["empresa"] == EMPRESA_FILTER].copy()
    desp_q1   = (
        xpto_desp.groupby("cod_conta")["valor"]
        .sum()
        .reset_index()
        .rename(columns={"valor": "rea_q1"})
    )
    for _, r in desp_q1.iterrows():
        cod = str(r["cod_conta"])
        if not cod.startswith("31"):  # never overwrite receita accounts (mart_rec_df has priority)
            realized[cod] = float(r["rea_q1"])

    # ── Walk Excel rows ───────────────────────────────────────────────────────
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb["DRE"]

    rows: list = []
    outline_of: dict = {}  # rows-list index → Excel outline level (true hierarchy depth)
    in_indicators_block = False  # flips True after the "Indicadores" header row

    for row_num, raw_row in enumerate(
        ws.iter_rows(min_row=7, max_row=ws.max_row, values_only=True), start=7
    ):
        if not any(v is not None for v in raw_row[:12]):
            continue
        cod      = raw_row[0]
        nome_raw = raw_row[2]
        if not nome_raw:
            continue

        nome = str(nome_raw).strip()
        ol   = ws.row_dimensions[row_num].outline_level
        idx  = len(rows)

        # Indicadores block: percent-valued KPIs at the bottom of the sheet.
        # Header row "Indicadores" itself has no values; subsequent rows
        # (COGS, SG&A, Margem Bruta, Margem EBITDA, Margem LL) hold decimal
        # ratios (0.5085 = 50,85 %) — must NOT be summed nor displayed as R$.
        if nome.lower() == "indicadores":
            in_indicators_block = True
            continue
        if in_indicators_block:
            q1_val = raw_row[10]
            if q1_val is None:
                continue  # blank trailing row (e.g. row 181 "Glosas" with no values)
            rows.append({
                "row_type":   "indicator",
                "level":      0,
                "is_receita": None,
                "is_percent": True,
                "nome":       nome,
                "cod_conta":  None,
                "orc_jan":    round(float(raw_row[4]) * 100, 2) if raw_row[4] is not None else None,
                "orc_fev":    round(float(raw_row[6]) * 100, 2) if raw_row[6] is not None else None,
                "orc_mar":    round(float(raw_row[8]) * 100, 2) if raw_row[8] is not None else None,
                "orc_q1":     round(float(q1_val) * 100, 2),
                "rea_q1":     None,   # filled later from margens
                "var_rs":     None,
                "var_pct":    None,
            })
            outline_of[idx] = 0
            continue

        if cod is not None:
            cod_str = str(int(cod))
            orc     = orc_lookup.get(cod_str, {"jan": 0, "fev": 0, "mar": 0, "q1": 0})
            rea_q1  = realized.get(cod_str, 0.0)
            orc_q1  = orc["q1"]
            var_rs  = rea_q1 - orc_q1
            var_pct = compute_variance_pct(rea_q1, orc_q1)
            outline_of[idx] = ol
            rows.append({
                "row_type":   "account",
                "level":      2,
                "is_receita": cod_str.startswith("31"),
                "nome":       nome,
                "cod_conta":  cod_str,
                "orc_jan":    round(orc["jan"], 2),
                "orc_fev":    round(orc["fev"], 2),
                "orc_mar":    round(orc["mar"], 2),
                "orc_q1":     round(orc_q1, 2),
                "rea_q1":     round(rea_q1, 2),
                "var_rs":     round(var_rs, 2),
                "var_pct":    var_pct,
            })
        else:
            # Section row: use raw Excel values (subtotals from formulas)
            jan   = float(raw_row[4]) if raw_row[4] is not None else 0.0
            fev   = float(raw_row[6]) if raw_row[6] is not None else 0.0
            mar   = float(raw_row[8]) if raw_row[8] is not None else 0.0
            orc_q1 = jan + fev + mar
            level  = 0 if nome in MAJOR_SECTIONS else 1
            outline_of[idx] = ol
            rows.append({
                "row_type":   "section",
                "level":      level,
                "is_receita": None,
                "nome":       nome,
                "cod_conta":  None,
                "orc_jan":    round(jan, 2) if jan else None,
                "orc_fev":    round(fev, 2) if fev else None,
                "orc_mar":    round(mar, 2) if mar else None,
                "orc_q1":     round(orc_q1, 2) if orc_q1 else None,
                "rea_q1":     None,
                "var_rs":     None,
                "var_pct":    None,
                "_outline":   ol,
            })

    # ── Inject known-unbudgeted accounts absent from the Excel DRE ───────────
    # These accounts have realized expenses but no budget line in the source
    # sheet. staging injects them (orcado=0) so the mart tracks them; we mirror
    # that injection here so the hierarchy stays consistent with all_contas
    # and the validation_gate reconciliation does not flag their value.
    existing_cods = {r["cod_conta"] for r in rows if r.get("cod_conta")}
    for cod, nome in KNOWN_UNBUDGETED_ACCOUNTS.items():
        if cod in existing_cods:
            continue
        rea_q1 = realized.get(cod, 0.0)
        rows.append({
            "row_type":   "account",
            "level":      2,
            "is_receita": False,
            "nome":       nome,
            "cod_conta":  cod,
            "orc_jan":    0.0, "orc_fev": 0.0, "orc_mar": 0.0, "orc_q1": 0.0,
            "rea_q1":     round(rea_q1, 2),
            "var_rs":     round(rea_q1, 2),
            "var_pct":    None,
        })
        outline_of[len(rows) - 1] = 2

    # ── Bottom-up pass: reconstruct missing January for subtotal section rows ──
    # Many Excel SUM formulas for January were not cached in the xlsx file.
    # We use outline_level (true hierarchy depth) to collect only direct
    # children (outline == parent + 1).
    # When both sub-section rows and leaf-account rows exist as direct children,
    # the sub-section rows already represent the same total as the accounts
    # (the accounts are the GL entries, the sections are display breakdowns).
    # Prefer section children to avoid double-counting.
    for i in range(len(rows) - 1, -1, -1):
        row = rows[i]
        if row["row_type"] != "section" or row["orc_jan"] is not None:
            continue
        ol_parent   = outline_of[i]
        ol_children = ol_parent + 1
        sec_vals: list = []
        acc_vals: list = []
        for j in range(i + 1, len(rows)):
            if outline_of[j] <= ol_parent:
                break
            if outline_of[j] == ol_children and rows[j]["orc_jan"] is not None:
                if rows[j]["row_type"] == "section":
                    sec_vals.append(rows[j]["orc_jan"])
                else:
                    acc_vals.append(rows[j]["orc_jan"])
        chosen = sec_vals if sec_vals else acc_vals
        if chosen:
            jan_sum = sum(chosen)
            fev = row["orc_fev"] or 0.0
            mar = row["orc_mar"] or 0.0
            row["orc_jan"] = round(jan_sum, 2)
            q1 = jan_sum + fev + mar
            row["orc_q1"]  = round(q1, 2) if q1 else None

    # ── Formula pass: fix computed rows that have no children ─────────────────
    # Receita Líquida, Lucro Bruto, EBITDA, Lucro Operacional, Lucro Líquido
    # are derived from surrounding sections, not from summing child accounts.
    by_nome = {r["nome"]: r for r in rows}

    def _gj(nome: str, default: float = 0.0) -> float:
        r = by_nome.get(nome)
        return r["orc_jan"] if (r and r["orc_jan"] is not None) else default

    def _fix(nome: str, jan_val):
        r = by_nome.get(nome)
        if r and r["orc_jan"] is None and jan_val is not None:
            r["orc_jan"] = round(jan_val, 2)
            fev = r["orc_fev"] or 0.0
            mar = r["orc_mar"] or 0.0
            q1  = jan_val + fev + mar
            r["orc_q1"] = round(q1, 2) if q1 else None

    rb   = _gj("Receita Bruta")
    ab   = _gj("Abatimentos")
    _fix("Receita Líquida", rb - ab if rb else None)

    rl   = _gj("Receita Líquida")
    cst  = _gj("Custo dos Serviços Prestados")
    _fix("Lucro Bruto", rl - cst if rl else None)

    lb   = _gj("Lucro Bruto")
    dop  = _gj("Despesas Operacionais")
    dap  = _gj("Depreciação E Amortização (Produção)")
    _fix("EBITDA", lb - dop + dap if lb else None)

    ebt  = _gj("EBITDA")
    da   = _gj("Depreciação E Amortização")
    _fix("Lucro Operacional", ebt - da if ebt else None)

    lo   = _gj("Lucro Operacional")
    rf   = _gj("Resultado Financeiro")
    ir   = _gj("Imposto De Renda e CSLL")
    ps   = _gj("Participação Sobre Os Resultados")
    _fix("Lucro Líquido", lo - rf - ir - ps if lo else None)

    # ── Realized pass: aggregate child accounts into section subtotals ───────
    # For each section row, sum rea_q1 of all descendant account rows
    # (rows below it until another same-or-higher-level section).
    for i, row in enumerate(rows):
        if row["row_type"] != "section":
            continue
        s_ol = row.get("_outline", 0)
        rea_sum = 0.0
        found = False
        for j in range(i + 1, len(rows)):
            other = rows[j]
            other_ol = other.get("_outline", 0)
            if other["row_type"] == "section" and other_ol <= s_ol:
                break
            if other["row_type"] == "account" and other.get("rea_q1") is not None:
                rea_sum += float(other["rea_q1"])
                found = True
        if found:
            row["rea_q1"] = round(rea_sum, 2)
            if row.get("orc_q1") is not None:
                row["var_rs"]  = round(rea_sum - row["orc_q1"], 2)
                row["var_pct"] = compute_variance_pct(rea_sum, row["orc_q1"])

    # ── Realized pass for derived sections (formulas, no children) ───────────
    def _gj_rea(nome: str, default: float = 0.0) -> float:
        r = by_nome.get(nome)
        return r["rea_q1"] if (r and r.get("rea_q1") is not None) else default

    def _fix_rea(nome: str, val):
        # Overwrite even if a previous pass populated rea_q1 — derived
        # sections (Receita Líquida, EBITDA, Lucro Líquido) must reflect
        # the inter-section formula, not a stray child sum (e.g. injected
        # KNOWN_UNBUDGETED accounts appended after Lucro Líquido).
        r = by_nome.get(nome)
        if r and val is not None:
            r["rea_q1"] = round(val, 2)
            if r.get("orc_q1") is not None:
                r["var_rs"]  = round(val - r["orc_q1"], 2)
                r["var_pct"] = compute_variance_pct(val, r["orc_q1"])

    rb_r  = _gj_rea("Receita Bruta")
    ab_r  = _gj_rea("Abatimentos")
    _fix_rea("Receita Líquida", rb_r - ab_r if rb_r else None)

    rl_r  = _gj_rea("Receita Líquida")
    cst_r = _gj_rea("Custo dos Serviços Prestados")
    _fix_rea("Lucro Bruto", rl_r - cst_r if rl_r else None)

    lb_r  = _gj_rea("Lucro Bruto")
    dop_r = _gj_rea("Despesas Operacionais")
    dap_r = _gj_rea("Depreciação E Amortização (Produção)")
    _fix_rea("EBITDA", lb_r - dop_r + dap_r if lb_r else None)

    ebt_r = _gj_rea("EBITDA")
    da_r  = _gj_rea("Depreciação E Amortização")
    _fix_rea("Lucro Operacional", ebt_r - da_r if ebt_r else None)

    lo_r  = _gj_rea("Lucro Operacional")
    rf_r  = _gj_rea("Resultado Financeiro")
    ir_r  = _gj_rea("Imposto De Renda e CSLL")
    ps_r  = _gj_rea("Participação Sobre Os Resultados")
    _fix_rea("Lucro Líquido", lo_r - rf_r - ir_r - ps_r if lo_r else None)

    # ── Indicadores: fill realizado% from already-computed sections ──────────
    # All ratios use Receita Líquida (ROL) as denominator.
    rl_r_ind  = _gj_rea("Receita Líquida")
    if rl_r_ind:
        cst_ind   = _gj_rea("Custo dos Serviços Prestados")
        dop_ind   = _gj_rea("Despesas Operacionais")
        lb_ind    = _gj_rea("Lucro Bruto")
        eb_ind    = _gj_rea("EBITDA")
        ll_ind    = _gj_rea("Lucro Líquido")

        ind_real_pct = {
            "COGS":                    cst_ind / rl_r_ind * 100 if cst_ind else 0,
            "SG&A":                    dop_ind / rl_r_ind * 100 if dop_ind else 0,
            "Margem Bruta (Rol)":      lb_ind  / rl_r_ind * 100 if lb_ind  else 0,
            "Margem EBITDA":           eb_ind  / rl_r_ind * 100 if eb_ind  else 0,
            "Margem Lucro Líquido (rol)": ll_ind / rl_r_ind * 100 if ll_ind else 0,
        }
        for r in rows:
            if r["row_type"] == "indicator" and r["nome"] in ind_real_pct:
                rea_pct = round(ind_real_pct[r["nome"]], 2)
                r["rea_q1"]  = rea_pct
                if r["orc_q1"] is not None:
                    r["var_rs"]  = round(rea_pct - r["orc_q1"], 2)  # delta in p.p.
                    r["var_pct"] = round((rea_pct - r["orc_q1"]) / abs(r["orc_q1"]) * 100, 2) if r["orc_q1"] else None

    # Strip internal _outline marker before returning
    for r in rows:
        r.pop("_outline", None)

    return rows


def _clean(obj):
    """Recursively replace float NaN/Infinity with None for valid JSON."""
    import math
    if isinstance(obj, float) and (math.isnan(obj) or math.isinf(obj)):
        return None
    if isinstance(obj, dict):
        return {k: _clean(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_clean(v) for v in obj]
    return obj


def main() -> None:
    OUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    logger.info("Building dashboard data...")
    data = build()   # _clean() is applied inside build()
    with open(OUT_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    logger.info("Saved: %s", OUT_FILE)
    logger.info("Summary: %s", data["summary"])


if __name__ == "__main__":
    main()
